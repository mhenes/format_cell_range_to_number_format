#!/usr/bin/env python
# Format range of cells to number format in excel.
# Mark Henes
import os
# openpyxl module imports
from openpyxl import Workbook
from openpyxl import load_workbook
# pyexcel module imports
import pyexcel as pe
import pyexcel.ext.xls # import it to handle xls file
import pyexcel.ext.xlsx # import it to handle xlsx file

# Script path to the directory it is located.
pcwd=os.path.dirname(os.path.abspath(__file__))

# List files in directory by file extension
# Specify directory
items = os.listdir(pcwd)

# Specify extension in "if" loop
worksheet_list = []
for names in items:
    if names.endswith(".xlsx"):
        worksheet_list.append(names)

# Format specified columns to '0' number format
for i in range(len(worksheet_list)):
    wb = load_workbook(filename = worksheet_list[i])
    ws = wb.active
    for row in ws.iter_rows('B2:C200'):
        for cell in row:
            cell.number_format = '0'
    wb.save(worksheet_list[i])
